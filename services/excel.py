import json
from datetime import datetime
from pathlib import Path
from typing import Any, Optional

from openpyxl import Workbook, load_workbook

import config


def _normalize_reg_no(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().upper()


def record_key(row: dict[str, Any]) -> tuple[str, str]:
    reg = _normalize_reg_no(row.get("registration_no"))
    if reg:
        return ("reg", reg)
    return ("profile", str(row.get("profile_id", "")).strip())


def _ordered_columns(rows: list[dict[str, Any]]) -> list[str]:
    extra = set()
    for row in rows:
        extra.update(row.keys())
    ordered = [c for c in config.COLUMN_ORDER if c in extra]
    for col in sorted(extra):
        if col not in ordered:
            ordered.append(col)
    return ordered


def write_excel(rows: list[dict[str, Any]], path: Path) -> None:
    config.ensure_dirs()
    wb = Workbook()
    ws = wb.active
    ws.title = "Biodata"

    columns = _ordered_columns(rows)
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col, "") for col in columns])
    wb.save(path)


def read_excel(path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    records = []
    for values in rows_iter:
        record = {}
        for header, value in zip(headers, values):
            if header:
                record[header] = value if value is not None else ""
        if record:
            records.append(record)
    wb.close()
    return records


def list_exports() -> list[dict[str, Any]]:
    config.ensure_dirs()
    exports = []
    for path in sorted(config.EXPORTS_DIR.glob("*.xlsx"), reverse=True):
        summary_path = config.SUMMARIES_DIR / f"{path.stem}.json"
        summary = {}
        if summary_path.exists():
            with open(summary_path, encoding="utf-8") as f:
                summary = json.load(f)
        exports.append(
            {
                "filename": path.name,
                "type": "full" if path.name.startswith("full_") else "new_only",
                "size_bytes": path.stat().st_size,
                "modified": datetime.fromtimestamp(path.stat().st_mtime).isoformat(),
                **summary,
            }
        )
    return exports


def latest_full_export() -> Optional[Path]:
    config.ensure_dirs()
    files = sorted(config.EXPORTS_DIR.glob("full_*.xlsx"), reverse=True)
    return files[0] if files else None


def get_baseline_path(baseline_file: str = "") -> Optional[Path]:
    if baseline_file:
        path = config.EXPORTS_DIR / baseline_file
        if path.exists() and path.suffix == ".xlsx":
            return path
    return latest_full_export()


def find_new_records(
    current: list[dict[str, Any]],
    baseline: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    baseline_keys = {record_key(row) for row in baseline}
    return [row for row in current if record_key(row) not in baseline_keys]


def export_scrape_results(
    records: list[dict[str, Any]],
    baseline_file: str = "",
    pages_fetched: int = 0,
) -> dict[str, Any]:
    config.ensure_dirs()
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    full_path = config.EXPORTS_DIR / f"full_{timestamp}.xlsx"
    new_path = config.EXPORTS_DIR / f"new_only_{timestamp}.xlsx"

    baseline_path = get_baseline_path(baseline_file)
    baseline_records = read_excel(baseline_path) if baseline_path else []
    new_records = find_new_records(records, baseline_records)

    write_excel(records, full_path)
    write_excel(new_records, new_path)

    summary = {
        "timestamp": timestamp,
        "total_scraped": len(records),
        "baseline_count": len(baseline_records),
        "baseline_file": baseline_path.name if baseline_path else None,
        "new_count": len(new_records),
        "pages_fetched": pages_fetched,
        "full_export": full_path.name,
        "new_only_export": new_path.name,
    }

    summary_path = config.SUMMARIES_DIR / f"full_{timestamp}.json"
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump(summary, f, indent=2)

    return summary
